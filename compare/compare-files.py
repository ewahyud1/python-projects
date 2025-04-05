from openpyxl import load_workbook
import requests
import argparse

"""
  This script is used to compare two lists of servers:
  List 1 (octo_list): Interfacing with Octopus REST API to harvest list of deployment targets (servers) in Cctopus Deploy
  List 2 (external_list): Excel file containing a list of servers to be compared with the Octopus list
  The script will return the intersection of the two lists, i.e., the servers that are present in both lists.
  The script uses the openpyxl library to read the Excel file and the requests library to make API calls.
  To run the script, execute the following command:
      python compare-files.py --apikey XXXXXXXXXX
"""

def compare_server_lists(external_list, octo_list):
   
    external_servers = []
    try:
        # Load the Excel workbook
        workbook = load_workbook(external_list)
        # Get the active spreadsheet
        sheet = workbook.active
        
        # Reading server names in column A, starting from row 1
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
            server_name = row[0].value
            if server_name:  # Check if cell is not empty
                external_servers.append(str(server_name).strip().lower())
                
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None        

    # Convert to sets for efficient comparison
    external_servers_set = set(external_servers)
    octo_servers_set = set(octo_list)

    # Perform Comparisons to find servers in both lists (intersection)
    results = (list(octo_servers_set.intersection(external_servers_set)))
    return results

### Using Octopus REST API to harvest deployment targets (aka servers)
def get_octopus_servers(api_key):
    #Octopus URL
    octopus_url = 'https://example.octopus.com'

    # Octopus API endpoint to get all servers
    endpoint = f'{octopus_url}/api/machines/all'

    # Headers for authentication
    headers = {
        'X-Octopus-ApiKey': api_key
    }

    # Create empty Array/List to store server names
    server_names = []

    # Make the GET request
    response = requests.get(endpoint, headers=headers)

    # Check if the request was successful
    if response.status_code == 200:
        servers = response.json()
        for server in servers:
            # Includes all enabled servers only            
            if not (server['IsDisabled']):
                server_names.append((server['Name']).lower())
        return server_names                           
        
    else:
        print(f"Failed to retrieve servers: {response.status_code}")
        return []

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Get Octopus servers list')
    parser.add_argument('--apikey', required=True, help='Octopus API token')
    args = parser.parse_args()
        
    #Path to  server list 1
    listOne = './<path-to>/servers-list-01.xlsx'

    #Path to server list 2
    listTwo = './<path-to>/servers-list-02.xlsx'

    # You can add a 3rd or more lists to compare
  
    #Compile Octopus Server List
    octopus_server_list = get_octopus_servers(args.apikey)

    # use compare_server_lists function to compare the lists
    listOneIntersection = compare_server_lists(listOne, octopus_server_list)
    listTwoIntersection = compare_server_lists(listTwo, octopus_server_list)
    
    print(f"\nServers list #1 comparison: {listOneIntersection}")
    print(f"\nServers list #2 comparison: {listTwoIntersection}")
